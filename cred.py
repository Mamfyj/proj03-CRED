import openpyxl as xl
import re as re
import os

arquivo = xl.load_workbook(filename="C:/Users/mamfy/Desktop/prog/proj03-CRED/exemplo_dados_com_cpf.xlsx")
output = open("output.txt", "x", encoding="utf-8")

ws = arquivo.active

def addPresenca():

    if ws.cell(row=1, column=ws.max_column).value != "PRESENÇA":
        ws.insert_cols(ws.max_column + 1)
        ws.cell(row=1,column=ws.max_column+1,value="PRESENÇA")

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            row[-1].value = "AUSENTE"
    
    arquivo.save("C:/Users/mamfy/Desktop/prog/proj03-CRED/exemplo_dados_com_cpf_teste.xlsx")

addPresenca()

def pesquisa(query, r):

    query = re.sub(r'\D', '', query)

    for row in ws.iter_rows(values_only=False): 
        if re.sub(r'\D', '', str(row[r].value)) == query:
            row[-1].value = "PRESENTE"
            arquivo.save("C:/Users/mamfy/Desktop/prog/proj03-CRED/exemplo_dados_com_cpf_teste.xlsx")
            print([cell.value for cell in row])
            output.write(",".join(str(cell.value) for cell in row) + "\n")
            return
    print("404")

while True:
    try:
        query = input("Digite o CPF / RG para pesquisa: ")
        if len(query) < 11:
            pesquisa(query, 1)
        else: pesquisa(query, 0)

        if query.lower() == 'sair':
            break
    except EOFError:
        break

